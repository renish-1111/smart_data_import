# Copyright (c) 2026, ERPNext AI Team and contributors
# For license information, please see license.txt

import csv
import gc
import io
import json
import os
import re
import time
from collections import defaultdict, deque

import openpyxl
from openpyxl.styles import Font, PatternFill

import frappe
from frappe import _
from frappe.utils import cint, flt, get_files_path

EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xltx", ".xltm")
CSV_EXTENSIONS = (".csv",)
SUPPORTED_EXTENSIONS = EXCEL_EXTENSIONS + CSV_EXTENSIONS


class ImportCancelled(Exception):
	"""Raised internally when a user-requested cancellation is detected between
	batches. Caught inside execute_import — never a real failure."""


# The engine writes with ignore_permissions=True, so it is the only thing standing
# between a caller and any DocType in the system. System Managers use this
# intentionally for trusted bulk loads. Anyone else (e.g. the self-service import
# portal) is additionally checked against their real create/write permission on the
# target DocType in _assert_target_doctypes_allowed — but these stay blocked even if
# a role was ever accidentally granted permission on them.
SECURITY_CRITICAL_DOCTYPES = {
	"User",
	"Role",
	"Role Profile",
	"Has Role",
	"DocType",
	"DocField",
	"DocPerm",
	"Custom DocPerm",
	"Custom Field",
	"Property Setter",
	"Module Def",
	"System Settings",
	"Server Script",
	"Client Script",
	"Webhook",
	"OAuth Client",
	"Integration Request",
	"Smart Data Import",
}

# Layout-only fieldtypes never hold data and must not take part in column mapping.
# How many problem rows are kept on the document for the UI. The full list always
# lives in the attached Excel log.
ERROR_ROWS_IN_UI = 500

LAYOUT_FIELDTYPES = {
	"Section Break",
	"Column Break",
	"Tab Break",
	"HTML",
	"Button",
	"Fold",
	"Heading",
	"Image",
}


def normalize_key(value):
	"""Normalizes a header/label/fieldname so `Customer Name`, `customer_name`
	and `CUSTOMER  NAME` all collapse to the same lookup key."""
	text = str(value or "").strip().lower().replace("_", " ").replace("-", " ")
	return re.sub(r"\s+", " ", text)


def resolve_file_path(file_url):
	"""Resolves an attached file URL to an absolute path on disk.

	Returns (path, error_message). Exactly one of the two is set.
	"""
	if not file_url:
		return None, _("No file attached in this row.")

	extension = os.path.splitext(file_url.split("?")[0])[1].lower()
	if extension == ".xls":
		return None, _("Legacy .xls files are not supported. Please re-save the file as .xlsx and attach it again.")
	if extension not in SUPPORTED_EXTENSIONS:
		return None, _("Unsupported file type {0}. Attach an Excel (.xlsx, .xlsm) or .csv file.").format(
			extension or file_url
		)

	file_path = None
	file_id = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if file_id:
		file_path = frappe.get_doc("File", file_id).get_full_path()
	else:
		site_relative = file_url.lstrip("/")
		file_path = frappe.get_site_path("public", site_relative)
		if not os.path.exists(file_path):
			file_path = frappe.get_site_path(site_relative)

	if not file_path or not os.path.exists(file_path):
		return None, _("File could not be found on disk. Please remove the row and upload the file again.")

	return file_path, None


def build_field_lookup(meta):
	"""Builds {normalized header -> fieldname} for a DocType, covering both
	field names and human readable labels."""
	lookup = {}
	for field in meta.fields:
		if field.fieldtype in LAYOUT_FIELDTYPES:
			continue
		lookup[normalize_key(field.fieldname)] = field.fieldname
		if field.label:
			lookup.setdefault(normalize_key(field.label), field.fieldname)

	# `ID` is the label Frappe itself uses for the primary key in exports.
	lookup["name"] = "name"
	lookup["id"] = "name"
	return lookup


def map_columns(meta, headers, column_overrides=None):
	"""Maps spreadsheet headers to DocType fieldnames.

	Returns a dict with:
	        col_map  - positional list (fieldname or None) used while streaming rows
	        mapped   - {header: fieldname}
	        unmapped - [header, ...] columns that will be ignored
	"""
	lookup = build_field_lookup(meta)
	overrides = {normalize_key(k): v for k, v in (column_overrides or {}).items()}

	col_map = []
	mapped = {}
	unmapped = []

	for header in headers:
		header_text = str(header).strip() if header is not None else ""
		key = normalize_key(header_text)
		fieldname = overrides.get(key) or lookup.get(key)

		col_map.append(fieldname)
		if fieldname:
			mapped[header_text] = fieldname
		elif header_text:
			unmapped.append(header_text)

	return {"col_map": col_map, "mapped": mapped, "unmapped": unmapped}


def missing_mandatory_fields(meta, mapped_fieldnames, defaults=None):
	"""Returns labels of mandatory fields that are neither in the file nor in the
	configured default values, so the user is warned *before* the import runs."""
	provided = set(mapped_fieldnames) | set((defaults or {}).keys())
	missing = []
	for field in meta.fields:
		if not field.reqd or field.fieldtype in LAYOUT_FIELDTYPES:
			continue
		if field.default:
			continue
		if field.fieldname in provided:
			continue
		missing.append(field.label or field.fieldname)
	return missing


def mandatory_fieldnames(meta):
	"""Fieldnames that are mandatory on this DocType, for highlighting them in the
	column mapping preview regardless of whether the file actually supplies them."""
	return {f.fieldname for f in meta.fields if f.reqd and f.fieldtype not in LAYOUT_FIELDTYPES}


def sniff_csv_dialect(file_path):
	"""Detects the delimiter so semicolon/tab separated exports work out of the box."""
	try:
		with open(file_path, encoding="utf-8-sig", errors="ignore") as f:
			sample = f.read(8192)
		return csv.Sniffer().sniff(sample, delimiters=",;\t|")
	except Exception:
		return csv.excel


class SmartImportEngine:
	"""
	High-performance batch data import engine.
	Handles automated dependency graph resolution (DAG), self-referential hierarchy
	sorting, memory-efficient streaming and chunked batch insertion.
	"""

	def __init__(self, import_doc):
		if isinstance(import_doc, str):
			self.doc = frappe.get_doc("Smart Data Import", import_doc)
		else:
			self.doc = import_doc

		self.batch_size = max(500, cint(self.doc.batch_size) or 5000)
		self.skip_empty_rows = bool(self.doc.skip_empty_rows)
		self.clean_whitespace = bool(self.doc.clean_whitespace)
		self.ignore_link_errors = bool(self.doc.ignore_link_errors)
		self.ignore_mandatory = bool(getattr(self.doc, "ignore_mandatory_errors", False))
		self.stop_on_error = bool(self.doc.stop_on_error)
		self.ignore_duplicates = bool(getattr(self.doc, "ignore_duplicates", False))
		self.import_type = self.doc.import_type or "Insert New Records"
		self.rules = self._parse_filter_rules()
		self._dt_index = None
		self._manifest_buffer = []
		self._link_resolve_cache = {}

	# -------------------------------------------------------------------------
	# 0. CONFIGURATION HELPERS
	# -------------------------------------------------------------------------

	def _parse_filter_rules(self):
		"""Parses the optional per-DocType mapping/defaults configuration:

		{
		    "Customer": {
		        "column_map": {"Cust Name": "customer_name"},
		        "defaults": {"customer_type": "Company"}
		    }
		}
		"""
		if not self.doc.filter_rules_json:
			return {}
		try:
			rules = json.loads(self.doc.filter_rules_json)
			return rules if isinstance(rules, dict) else {}
		except Exception:
			return {}

	def rules_for(self, target_doctype):
		rules = self.rules.get(target_doctype) or {}
		if not isinstance(rules, dict):
			return {}, {}
		column_map = rules.get("column_map") or {}
		defaults = rules.get("defaults") or {}
		return (
			column_map if isinstance(column_map, dict) else {},
			defaults if isinstance(defaults, dict) else {},
		)

	def supports(self, fieldname):
		"""True when the site's schema actually has this field.

		Code can reach a site before `bench migrate` does; when that happens the newer
		features degrade quietly instead of crashing a running import.
		"""
		return bool(self.doc.meta.has_field(fieldname))

	def supported_updates(self, values):
		"""Filters an update dict down to columns that exist on this site."""
		return {k: v for k, v in values.items() if self.supports(k)}

	@property
	def processed_records(self):
		return (
			cint(self.doc.imported_records)
			+ cint(self.doc.failed_records)
			+ cint(getattr(self.doc, "skipped_records", 0))
		)

	# -------------------------------------------------------------------------
	# 1. FILE ANALYSIS & DOCTYPE AUTO-DETECTION
	# -------------------------------------------------------------------------

	def analyze_files_and_build_graph(self):
		"""
		Reads attached files, infers target DocTypes, counts rows, validates the
		column mapping and builds the DAG used for dependency resolution.
		"""
		self.doc.status = "Analyzing"
		self.doc.save(ignore_permissions=True)
		frappe.db.commit()

		if getattr(self.doc, "expand_multi_sheet_files", 1):
			self._expand_multi_sheet_files()

		file_sheet_counts = defaultdict(int)
		for row in self.doc.files:
			if row.file:
				file_sheet_counts[row.file] += 1

		detected_doctypes = set()
		total_files = len(self.doc.files) or 1

		for file_idx, row in enumerate(self.doc.files, start=1):
			row.mapping_summary = ""
			row.error_log = ""
			self.publish_progress(
				_("Analyzing file {0} of {1}...").format(file_idx, total_files),
				flt((file_idx - 1) / total_files * 100, 1),
			)

			file_path, error = resolve_file_path(row.file)
			if error:
				row.status = "Failed"
				row.error_log = error
				row.total_rows = 0
				continue

			try:
				headers, row_count, sheet_name = self.read_file_header_and_count(file_path, row.sheet_name)
			except Exception as e:
				row.status = "Failed"
				row.error_log = _("Could not read file: {0}").format(str(e))
				row.total_rows = 0
				continue

			row.total_rows = row_count
			if sheet_name:
				row.sheet_name = sheet_name

			# Auto-detection only fills blanks; a manually chosen DocType is never overwritten.
			if not row.doctype_name and self.doc.auto_detect_doctype:
				is_multi_sheet = file_sheet_counts.get(row.file, 0) > 1
				detected, note = self.detect_target_doctype(
					row.file, row.sheet_name, headers, is_multi_sheet=is_multi_sheet
				)
				if detected:
					row.doctype_name = detected
					row.mapping_summary = note

			if not row.doctype_name:
				row.status = "Failed"
				row.error_log = _(
					"Could not detect the target DocType. Please pick it manually in the 'Target DocType' column."
				)
				continue

			if not frappe.db.exists("DocType", row.doctype_name):
				row.status = "Failed"
				row.error_log = _("DocType {0} does not exist.").format(row.doctype_name)
				continue

			row.status = "Analyzed"
			detected_doctypes.add(row.doctype_name)

			if not row_count:
				row.status = "Failed"
				row.error_log = _("No data rows found. Row 1 must contain the column headers.")
				continue

			row.mapping_summary = self._describe_mapping(row.doctype_name, headers, row.mapping_summary)

		self.publish_progress(_("Building dependency graph..."), 100)

		# Build dependency graph & topological tiers
		self.doc.dependencies = []
		total_records = 0
		if detected_doctypes:
			for dt_info in self.build_topological_dependency_tiers(detected_doctypes):
				total_dt_rows = sum(
					cint(r.total_rows) for r in self.doc.files if r.doctype_name == dt_info["doctype"]
				)
				total_records += total_dt_rows
				self.doc.append(
					"dependencies",
					{
						"execution_tier": dt_info["tier"],
						"doctype_name": dt_info["doctype"],
						"depends_on_doctypes": ", ".join(dt_info["depends_on"]),
						"has_inner_dependency": 1 if dt_info["inner_ref_field"] else 0,
						"self_reference_field": dt_info["inner_ref_field"] or "",
						"status": "Pending",
						"total_count": total_dt_rows,
						"processed_count": 0,
					},
				)

		self.doc.total_records = total_records
		self.doc.status = "Ready" if total_records else "Pending"
		self.doc.save(ignore_permissions=True)
		frappe.db.commit()
		return True

	def _expand_multi_sheet_files(self):
		"""Auto-adds one Files row per extra sheet in an uploaded workbook.

		A user attaching a single .xlsx that has several sheets (e.g. one per
		DocType) only ever got the first sheet imported, since one row maps to
		one sheet. This mirrors that automatically so every page of the
		workbook ends up as its own row, each analyzed and auto-detected
		independently. Idempotent: re-analyzing never creates duplicate rows
		for a sheet that already has one, and the "Field Guide" tab this same
		app writes into its own downloadable templates is always skipped.
		"""
		existing_by_file = defaultdict(set)
		for row in self.doc.files:
			existing_by_file[row.file].add(row.sheet_name or "")

		new_rows = []
		for row in list(self.doc.files):
			if not row.file:
				continue

			file_path, error = resolve_file_path(row.file)
			if error or not file_path.lower().endswith(EXCEL_EXTENSIONS):
				continue

			try:
				wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
				sheet_names = list(wb.sheetnames)
				wb.close()
			except Exception:
				continue

			if len(sheet_names) <= 1:
				continue

			if not row.sheet_name:
				row.sheet_name = sheet_names[0]
				existing_by_file[row.file].add(row.sheet_name)

			for extra_sheet in sheet_names:
				if extra_sheet == "Field Guide" or extra_sheet in existing_by_file[row.file]:
					continue
				try:
					_headers, count, _sn = self.read_file_header_and_count(file_path, extra_sheet)
				except Exception:
					continue
				if not count:
					continue
				existing_by_file[row.file].add(extra_sheet)
				new_rows.append({"file": row.file, "sheet_name": extra_sheet})

		for new_row in new_rows:
			self.doc.append("files", new_row)

	def _describe_mapping(self, target_doctype, headers, prefix=""):
		"""Builds the short human readable mapping report shown on each file row."""
		meta = frappe.get_meta(target_doctype)
		column_map, defaults = self.rules_for(target_doctype)
		mapping = map_columns(meta, headers, column_map)
		missing = missing_mandatory_fields(meta, mapping["mapped"].values(), defaults)

		notes = [prefix] if prefix else []
		notes.append(_("{0} of {1} columns mapped.").format(len(mapping["mapped"]), len(mapping["mapped"]) + len(mapping["unmapped"])))
		if mapping["unmapped"]:
			notes.append(_("Ignored columns: {0}").format(", ".join(mapping["unmapped"][:10])))
		if missing:
			notes.append(_("⚠ Mandatory fields not in file: {0}").format(", ".join(missing[:10])))
		return " ".join(notes)

	def read_file_header_and_count(self, file_path, sheet_name=None):
		"""
		Reads headers and counts data rows without loading the whole file into RAM.
		Empty header cells keep their position so column alignment is preserved.
		"""
		headers = []
		count = 0
		selected_sheet = sheet_name

		if file_path.lower().endswith(EXCEL_EXTENSIONS):
			wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
			try:
				sheet_names = wb.sheetnames
				selected_sheet = sheet_name if (sheet_name and sheet_name in sheet_names) else sheet_names[0]
				ws = wb[selected_sheet]
				for idx, row in enumerate(ws.iter_rows(values_only=True)):
					if idx == 0:
						headers = self._clean_header_row(row)
					elif any(cell is not None and str(cell).strip() != "" for cell in row):
						count += 1
			finally:
				wb.close()
		elif file_path.lower().endswith(CSV_EXTENSIONS):
			dialect = sniff_csv_dialect(file_path)
			with open(file_path, encoding="utf-8-sig", errors="ignore", newline="") as f:
				for idx, row in enumerate(csv.reader(f, dialect)):
					if idx == 0:
						headers = self._clean_header_row(row)
					elif any(cell is not None and str(cell).strip() != "" for cell in row):
						count += 1

		return headers, count, selected_sheet

	@staticmethod
	def _clean_header_row(row):
		"""Keeps positional alignment (blank cells become ''), drops trailing blanks."""
		headers = ["" if cell is None else str(cell).strip() for cell in row]
		while headers and headers[-1] == "":
			headers.pop()
		return headers

	def _get_doctype_index(self):
		"""Builds a single in-memory index of {doctype: field keys} using two queries
		instead of calling frappe.get_meta() for every DocType on the site."""
		if self._dt_index is not None:
			return self._dt_index

		index = {}
		for dt in frappe.get_all("DocType", filters={"istable": 0, "issingle": 0}, fields=["name", "title_field"]):
			index[dt.name] = {"title_field": dt.title_field, "keys": set()}

		DocField = frappe.qb.DocType("DocField")
		rows = frappe.qb.from_(DocField).select(DocField.parent, DocField.fieldname, DocField.label).run(as_dict=True)

		CustomField = frappe.qb.DocType("Custom Field")
		rows += (
			frappe.qb.from_(CustomField)
			.select(CustomField.dt.as_("parent"), CustomField.fieldname, CustomField.label)
			.run(as_dict=True)
		)

		for row in rows:
			entry = index.get(row.parent)
			if not entry:
				continue
			if row.fieldname:
				entry["keys"].add(normalize_key(row.fieldname))
			if row.label:
				entry["keys"].add(normalize_key(row.label))

		self._dt_index = index
		return index

	def detect_target_doctype(self, filename, sheet_name, headers, is_multi_sheet=False):
		"""
		Infers the target DocType with a weighted score over header fields,
		primary-key naming, sheet name and filename word overlap.

		`is_multi_sheet` must be True when this row's file has more than one
		sheet. In that case every sheet in the workbook shares the same
		filename, so filename word overlap is not sheet-specific and is
		skipped — otherwise it can outweigh the correct header/sheet-name
		match for every sheet except the one the file happens to be named
		after.

		Returns (doctype, note) where note explains the confidence to the user.
		"""
		index = self._get_doctype_index()
		scores = defaultdict(int)

		header_set = {normalize_key(h) for h in headers if h}
		header_set.discard("")

		# Step 1: header/label matching (highest confidence signal)
		if header_set:
			for dt, entry in index.items():
				matches = len(header_set & entry["keys"])
				if normalize_key(dt) + " name" in header_set:
					matches += 3
				# A column literally named after the DocType itself (e.g. a
				# "Designation" header for the Designation master) is strong,
				# doctype-specific evidence — unlike the field/label intersection
				# above, which also matches every other DocType that merely has a
				# Link field pointing at this one (Employee.designation, etc.) and
				# so can't tell them apart on a sparse, single-column sheet.
				if normalize_key(dt) in header_set:
					matches += 5
				if entry["title_field"] and normalize_key(entry["title_field"]) in header_set:
					matches += 1
				if matches >= 2:
					scores[dt] += matches * 10

		# Step 2: sheet name matching
		if sheet_name:
			clean_sheet = normalize_key(sheet_name)
			for dt in index:
				if normalize_key(dt) == clean_sheet:
					scores[dt] += 40

		# Step 3: filename matching & word overlap. Skipped for a multi-sheet
		# workbook, since every sheet in it shares this same filename.
		if not is_multi_sheet:
			base_name = os.path.basename(str(filename or "")).rsplit(".", 1)[0]
			clean_base = normalize_key(base_name)
			base_words = set(clean_base.split())
			for dt in index:
				dt_clean = normalize_key(dt)
				if dt_clean == clean_base:
					scores[dt] += 50
				else:
					common = set(dt_clean.split()) & base_words
					if common:
						scores[dt] += len(common) * 5

		if not scores:
			return None, ""

		ranked = sorted(scores.items(), key=lambda x: (-x[1], x[0]))
		best, best_score = ranked[0]
		note = _("Auto-detected {0}.").format(best)
		if len(ranked) > 1 and ranked[1][1] >= best_score * 0.9:
			note = _("Auto-detected {0} (close match with {1} — please confirm).").format(best, ranked[1][0])
		return best, note

	# -------------------------------------------------------------------------
	# 2. TOPOLOGICAL SORT FOR INTER & INTRA DOCTYPE DEPENDENCIES
	# -------------------------------------------------------------------------

	def build_topological_dependency_tiers(self, doctypes_set):
		"""
		Builds a DAG for the target DocTypes and runs Kahn's topological sort to
		determine execution tiers. Primary document Link fields take priority over
		child table links so cycles are avoided.
		"""
		doctypes = sorted(doctypes_set)
		graph = defaultdict(set)  # u -> nodes that depend on u
		in_degree = {dt: 0 for dt in doctypes}
		depends_on_map = defaultdict(set)
		inner_ref_map = {}

		for dt in doctypes:
			meta = frappe.get_meta(dt)
			inner_ref = None
			primary_link_doctypes = set()

			for f in meta.fields:
				if f.fieldtype == "Link" and f.options:
					if f.options == dt:
						inner_ref = f.fieldname
					elif f.options in doctypes_set:
						primary_link_doctypes.add(f.options)

			inner_ref_map[dt] = inner_ref
			depends_on_map[dt] = primary_link_doctypes

			for target_dt in primary_link_doctypes:
				if dt not in graph[target_dt]:
					graph[target_dt].add(dt)
					in_degree[dt] += 1

		queue = deque([dt for dt in doctypes if in_degree[dt] == 0])
		result_tiers = []
		current_tier = 0
		visited_count = 0

		while queue:
			for _i in range(len(queue)):
				u = queue.popleft()
				visited_count += 1
				result_tiers.append(
					{
						"doctype": u,
						"tier": current_tier,
						"depends_on": sorted(depends_on_map[u]),
						"inner_ref_field": inner_ref_map[u],
					}
				)
				for v in graph[u]:
					in_degree[v] -= 1
					if in_degree[v] == 0:
						queue.append(v)
			current_tier += 1

		# Cyclic leftovers go into a final tier so they are still imported.
		if visited_count < len(doctypes):
			for u in [dt for dt in doctypes if in_degree[dt] > 0]:
				result_tiers.append(
					{
						"doctype": u,
						"tier": current_tier,
						"depends_on": sorted(depends_on_map[u]),
						"inner_ref_field": inner_ref_map[u],
					}
				)

		return result_tiers

	# -------------------------------------------------------------------------
	# 3. HIGH-SPEED BATCH IMPORT EXECUTION ENGINE
	# -------------------------------------------------------------------------

	def execute_import(self):
		"""
		Executes the import tier by tier, streaming rows, flushing chunks and
		publishing real-time progress.
		"""
		start_time = time.time()
		self.doc.status = "Processing"
		self.doc.imported_records = 0
		self.doc.failed_records = 0
		self.doc.skipped_records = 0
		self.doc.progress_percent = 0
		self.doc.error_log = ""
		self.doc.save(ignore_permissions=True)
		frappe.db.commit()

		sorted_deps = sorted(self.doc.dependencies, key=lambda d: cint(d.execution_tier))
		row_log = []

		try:
			for dep in sorted_deps:
				self._raise_if_cancelled()
				dep.status = "Processing"
				self.doc.save(ignore_permissions=True)
				frappe.db.commit()

				target_dt = dep.doctype_name
				self.publish_progress(_("Tier {0}: importing {1}...").format(dep.execution_tier, target_dt))

				matching_files = [f for f in self.doc.files if f.doctype_name == target_dt]
				for file_row in matching_files:
					file_path, error = resolve_file_path(file_row.file)
					if error:
						file_row.status = "Failed"
						file_row.error_log = error
						row_log.append({"row": 0, "doctype": target_dt, "type": "Failed", "reason": error})
						continue

					file_row.status = "Importing"
					# Counters are advanced chunk-by-chunk inside _update_progress_counts,
					# so the totals returned here are used for reporting only.
					count, failed, skipped, errors = self._stream_and_batch_insert(
						file_path=file_path,
						sheet_name=file_row.sheet_name,
						target_doctype=target_dt,
						self_ref_field=dep.self_reference_field if dep.has_inner_dependency else None,
						dep_row=dep,
					)
					row_log.extend(errors)

					file_row.status = "Failed" if (failed and not count) else "Completed"
					file_row.error_log = _("{0} imported, {1} failed, {2} skipped.").format(count, failed, skipped)

					if failed and self.stop_on_error:
						dep.status = "Failed"
						self._finalize(start_time, row_log, "Failed")
						return False

				dep.status = "Completed" if cint(dep.processed_count) else "Pending"
				self.doc.save(ignore_permissions=True)
				frappe.db.commit()
		except ImportCancelled:
			# Whatever was imported before the cancel lands stays — no rollback,
			# consistent with how partial progress already works on a stop_on_error abort.
			self._finalize(start_time, row_log, "Cancelled")
			return False

		failed_only = [e for e in row_log if e.get("type") != "Skipped"]
		if failed_only:
			status = "Partial Success" if cint(self.doc.imported_records) else "Failed"
		else:
			status = "Completed"

		self._finalize(start_time, row_log, status)
		return True

	def _raise_if_cancelled(self):
		if frappe.db.get_value("Smart Data Import", self.doc.name, "status") == "Cancelled":
			raise ImportCancelled

	def _finalize(self, start_time, row_log, status):
		self._flush_manifest()
		self._attach_manifest()
		self._store_error_rows(row_log)
		self.doc.execution_time_seconds = flt(time.time() - start_time, 2)
		self.doc.status = status
		if row_log:
			self._generate_failed_rows_excel(row_log)
			self.doc.error_log = "\n".join(
				f"Row {entry['row']} [{entry.get('type', 'Failed')}]: {entry['reason']}" for entry in row_log[:200]
			)
			if len(row_log) > 200:
				self.doc.error_log += "\n" + _("... {0} more entries, see the downloadable log.").format(
					len(row_log) - 200
				)
		self.doc.progress_percent = 100
		self.doc.save(ignore_permissions=True)
		frappe.db.commit()
		self.publish_progress(
			_("Import finished: {0} imported, {1} failed, {2} skipped.").format(
				self.doc.imported_records, self.doc.failed_records, self.doc.skipped_records
			),
			100,
		)

	def _stream_and_batch_insert(self, file_path, sheet_name, target_doctype, self_ref_field, dep_row):
		"""
		Streams rows in configurable batch sizes and handles self-referential
		hierarchies with a two-pass strategy.
		"""
		meta = frappe.get_meta(target_doctype)
		column_map, defaults = self.rules_for(target_doctype)

		inserted_count = 0
		failed_count = 0
		skipped_count = 0
		row_log = []

		row_generator = self._get_row_generator(file_path, sheet_name)
		headers = next(row_generator, None)
		if not headers:
			return 0, 0, 0, []

		mapping = map_columns(meta, self._clean_header_row(headers), column_map)
		col_map = mapping["col_map"]
		if not mapping["mapped"]:
			reason = _("None of the columns in this file match fields of {0}.").format(target_doctype)
			self._add_counts(dep_row, 0, 1, 0)
			return 0, 1, 0, [{"row": 1, "doctype": target_doctype, "type": "Failed", "reason": reason}]

		title_field = self._get_title_field(meta, target_doctype)

		batch_records = []
		second_pass_updates = []
		row_idx = 1  # header occupies row 1

		for row_values in row_generator:
			row_idx += 1
			row_dict = dict(defaults)
			for idx, val in enumerate(row_values):
				fieldname = col_map[idx] if idx < len(col_map) else None
				if not fieldname or val is None:
					continue
				value = val.strip() if (self.clean_whitespace and isinstance(val, str)) else val
				if value != "":
					row_dict[fieldname] = value

			if self.skip_empty_rows and not any(k not in defaults for k in row_dict):
				continue

			row_dict = self._resolve_row_link_fields(target_doctype, row_dict, meta)

			# Pass 1 drops the self reference, pass 2 sets it once all rows exist.
			if self_ref_field and row_dict.get(self_ref_field):
				parent_val = row_dict.pop(self_ref_field)
				child_key = row_dict.get("name") or (row_dict.get(title_field) if title_field else None)
				if child_key:
					second_pass_updates.append((child_key, self_ref_field, parent_val))

			batch_records.append((row_idx, row_dict))

			if len(batch_records) >= self.batch_size:
				c, f, s, errs = self._flush_batch_to_db(target_doctype, batch_records, meta, title_field)
				inserted_count += c
				failed_count += f
				skipped_count += s
				row_log.extend(errs)
				batch_records = []
				self._update_progress_counts(dep_row, c, f, s)
				gc.collect()
				if f and self.stop_on_error:
					return inserted_count, failed_count, skipped_count, row_log
				self._raise_if_cancelled()

		if batch_records:
			c, f, s, errs = self._flush_batch_to_db(target_doctype, batch_records, meta, title_field)
			inserted_count += c
			failed_count += f
			skipped_count += s
			row_log.extend(errs)
			self._update_progress_counts(dep_row, c, f, s)
			self._raise_if_cancelled()

		if second_pass_updates:
			self.publish_progress(_("Linking hierarchy parents for {0}...").format(target_doctype))
			row_log.extend(self._apply_second_pass(target_doctype, meta, title_field, second_pass_updates))

		return inserted_count, failed_count, skipped_count, row_log

	@staticmethod
	def _get_title_field(meta, target_doctype):
		if meta.title_field:
			return meta.title_field
		for candidate in (
			target_doctype.lower().replace(" ", "_") + "_name",
			"title",
			"subject",
			"full_name",
		):
			if meta.has_field(candidate):
				return candidate
		return None

	def _apply_second_pass(self, target_doctype, meta, title_field, updates):
		"""Sets self-referential parent links now that every row exists.

		Tree DocTypes (Item Group, Customer Group, Account, Task, ...) are saved
		through the document so the nested set (lft/rgt) stays consistent — writing
		the column directly would leave the tree corrupted. Flat DocTypes take the
		fast direct-write path.
		"""
		is_tree = bool(getattr(meta, "is_tree", 0)) or (meta.has_field("lft") and meta.has_field("rgt"))
		problems = []

		for child_key, fieldname, parent_val in updates:
			try:
				child_name = child_key
				if not frappe.db.exists(target_doctype, child_name) and title_field:
					child_name = frappe.db.get_value(target_doctype, {title_field: child_key}, "name")

				parent_name = parent_val
				if not frappe.db.exists(target_doctype, parent_name) and title_field:
					parent_name = frappe.db.get_value(target_doctype, {title_field: parent_val}, "name")

				if not child_name or not parent_name:
					problems.append(
						{
							"row": 0,
							"doctype": target_doctype,
							"type": "Hierarchy",
							"reason": _("Could not link {0} to parent {1} — record not found.").format(
								child_key, parent_val
							),
						}
					)
					continue

				if is_tree:
					child = frappe.get_doc(target_doctype, child_name)
					child.set(fieldname, parent_name)
					child.flags.ignore_permissions = True
					child.save(ignore_permissions=True)
				else:
					frappe.db.set_value(
						target_doctype, child_name, fieldname, parent_name, update_modified=False
					)
			except Exception as e:
				problems.append(
					{
						"row": 0,
						"doctype": target_doctype,
						"type": "Hierarchy",
						"reason": _("Parent link {0} -> {1} failed: {2}").format(
							child_key, parent_val, self._clean_error(e)
						),
					}
				)

		frappe.db.commit()
		return problems

	def _resolve_row_link_fields(self, target_doctype, row_dict, meta):
		"""
		Resolves human readable Link titles to primary keys for autonamed DocTypes,
		e.g. Task.project = "Kintech AI Engine" -> "PROJ-0001".

		Results are cached per (link doctype, value) for the life of the import run,
		since the same Link value (e.g. a Territory or Customer Group) typically
		repeats across many rows — this turns what would be one query per row into
		one query per distinct value.
		"""
		resolved = dict(row_dict)
		for f in meta.fields:
			if f.fieldtype != "Link" or not f.options or f.fieldname not in resolved:
				continue
			val = resolved[f.fieldname]
			if not val or not isinstance(val, str):
				continue
			link_dt = f.options
			cache_key = (link_dt, val)

			if cache_key in self._link_resolve_cache:
				cached = self._link_resolve_cache[cache_key]
				if cached is not None:
					resolved[f.fieldname] = cached
				continue

			if frappe.db.exists(link_dt, val):
				self._link_resolve_cache[cache_key] = val
				continue

			link_meta = frappe.get_meta(link_dt)
			search_fields = []
			if link_meta.title_field:
				search_fields.append(link_meta.title_field)
			search_fields.append(link_dt.lower().replace(" ", "_") + "_name")
			search_fields.extend(["title", "subject", "item_name", "full_name"])

			found = None
			for sf in search_fields:
				if link_meta.has_field(sf):
					res = frappe.db.get_value(link_dt, {sf: val}, "name")
					if res:
						found = res
						break

			self._link_resolve_cache[cache_key] = found
			if found:
				resolved[f.fieldname] = found
		return resolved

	def _find_existing_name(self, target_doctype, meta, row_dict, title_field):
		"""Locates an existing record for duplicate detection / update modes."""
		name = row_dict.get("name")
		if name and frappe.db.exists(target_doctype, name):
			return name

		for f in meta.fields:
			if f.unique and row_dict.get(f.fieldname):
				existing = frappe.db.get_value(target_doctype, {f.fieldname: row_dict[f.fieldname]}, "name")
				if existing:
					return existing

		if title_field and row_dict.get(title_field):
			return frappe.db.get_value(target_doctype, {title_field: row_dict[title_field]}, "name")

		return None

	def _flush_batch_to_db(self, target_doctype, batch_records, meta, title_field):
		"""Writes one chunk of records, honouring the selected import mode."""
		c_success = 0
		c_failed = 0
		c_skipped = 0
		row_log = []

		needs_lookup = self.ignore_duplicates or self.import_type != "Insert New Records"

		for row_idx, row_dict in batch_records:
			try:
				existing = (
					self._find_existing_name(target_doctype, meta, row_dict, title_field)
					if needs_lookup
					else None
				)

				if existing and self.import_type in ("Update Existing Records", "Insert and Update"):
					self._update_existing(target_doctype, existing, row_dict)
					self._record_in_manifest("U", target_doctype, existing)
					c_success += 1
					continue

				if existing and self.ignore_duplicates:
					c_skipped += 1
					row_log.append(
						{
							"row": row_idx,
							"doctype": target_doctype,
							"type": "Skipped",
							"reason": _("Already exists as {0}").format(existing),
						}
					)
					continue

				if not existing and self.import_type == "Update Existing Records":
					c_skipped += 1
					row_log.append(
						{
							"row": row_idx,
							"doctype": target_doctype,
							"type": "Skipped",
							"reason": _("No matching record found to update."),
						}
					)
					continue

				doc = frappe.new_doc(target_doctype)
				doc.update(row_dict)
				doc.flags.ignore_permissions = True
				doc.insert(
					ignore_permissions=True,
					ignore_mandatory=self.ignore_mandatory,
					ignore_links=self.ignore_link_errors,
				)
				self._record_in_manifest("I", target_doctype, doc.name)
				c_success += 1

			except Exception as e:
				is_duplicate = isinstance(e, frappe.DuplicateEntryError) or "Duplicate" in type(e).__name__
				if is_duplicate and self.ignore_duplicates:
					c_skipped += 1
					row_log.append(
						{
							"row": row_idx,
							"doctype": target_doctype,
							"type": "Skipped",
							"reason": _("Duplicate entry skipped."),
						}
					)
				else:
					c_failed += 1
					row_log.append(
						{
							"row": row_idx,
							"doctype": target_doctype,
							"type": "Failed",
							"reason": self._clean_error(e),
						}
					)
					if self.stop_on_error:
						self._flush_manifest()
						frappe.db.commit()
						return c_success, c_failed, c_skipped, row_log

		self._flush_manifest()
		frappe.db.commit()
		return c_success, c_failed, c_skipped, row_log

	def _update_existing(self, target_doctype, name, row_dict):
		doc = frappe.get_doc(target_doctype, name)
		payload = {k: v for k, v in row_dict.items() if k != "name"}
		doc.update(payload)
		doc.flags.ignore_permissions = True
		doc.save(ignore_permissions=True)

	@staticmethod
	def _clean_error(exception):
		"""Turns Frappe's HTML-flavoured validation messages into a readable line."""
		message = str(exception) or type(exception).__name__
		message = re.sub(r"<[^>]+>", " ", message)
		message = re.sub(r"\s+", " ", message).strip()
		return message[:500]

	def _get_row_generator(self, file_path, sheet_name):
		"""Yields rows one at a time for Excel or CSV files."""
		if file_path.lower().endswith(EXCEL_EXTENSIONS):
			wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
			try:
				ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
				yield from ws.iter_rows(values_only=True)
			finally:
				wb.close()
		elif file_path.lower().endswith(CSV_EXTENSIONS):
			dialect = sniff_csv_dialect(file_path)
			with open(file_path, encoding="utf-8-sig", errors="ignore", newline="") as f:
				yield from csv.reader(f, dialect)

	def _add_counts(self, dep_row, inserted, failed, skipped):
		"""Single place where the running counters are advanced, so nothing is
		counted twice."""
		dep_row.processed_count = cint(dep_row.processed_count) + inserted + failed + skipped
		self.doc.imported_records = cint(self.doc.imported_records) + inserted
		self.doc.failed_records = cint(self.doc.failed_records) + failed
		self.doc.skipped_records = cint(self.doc.skipped_records) + skipped

	def _update_progress_counts(self, dep_row, inserted, failed, skipped):
		"""Publishes progress using lightweight writes so an open form is not
		flagged as modified while the background job runs."""
		self._add_counts(dep_row, inserted, failed, skipped)

		if cint(self.doc.total_records) > 0:
			self.doc.progress_percent = min(
				99.0, flt((self.processed_records / cint(self.doc.total_records)) * 100, 1)
			)

		frappe.db.set_value(
			"Smart Data Import",
			self.doc.name,
			self.supported_updates(
				{
					"imported_records": self.doc.imported_records,
					"failed_records": self.doc.failed_records,
					"skipped_records": self.doc.skipped_records,
					"progress_percent": self.doc.progress_percent,
				}
			),
			update_modified=False,
		)
		frappe.db.set_value(
			"Smart Data Import Dependency",
			dep_row.name,
			"processed_count",
			dep_row.processed_count,
			update_modified=False,
		)
		frappe.db.commit()
		self.publish_progress(
			_("{0}: {1} of {2} rows processed...").format(
				dep_row.doctype_name, dep_row.processed_count, dep_row.total_count
			)
		)

	def publish_progress(self, message, progress_pct=None):
		# A realtime/redis hiccup must never abort a long running import.
		try:
			self._publish(message, progress_pct)
		except Exception:
			pass

	def _publish(self, message, progress_pct=None):
		frappe.publish_realtime(
			"smart_import_progress",
			{
				"doc_name": self.doc.name,
				"message": message,
				"progress": progress_pct if progress_pct is not None else self.doc.progress_percent,
				"imported": cint(self.doc.imported_records),
				"failed": cint(self.doc.failed_records),
				"skipped": cint(getattr(self.doc, "skipped_records", 0)),
				"rolled_back": cint(getattr(self.doc, "rolled_back_records", 0)),
				"status": self.doc.status,
			},
			user=self.doc.owner,
			after_commit=False,
		)

	def _store_error_rows(self, row_log):
		"""Puts the problem rows on the document itself, so the reasons are readable
		in the form without downloading anything.

		Only the first ERROR_ROWS_IN_UI entries are stored — the complete list stays
		in the attached Excel log, which is what large imports need anyway.
		"""
		if not self.supports("errors"):
			return

		self.doc.errors = []
		for entry in row_log[:ERROR_ROWS_IN_UI]:
			self.doc.append(
				"errors",
				{
					"row_index": cint(entry.get("row")),
					"error_type": entry.get("type") or "Failed",
					"doctype_name": entry.get("doctype"),
					"reason": entry.get("reason"),
				},
			)

	def _generate_failed_rows_excel(self, row_log):
		"""Attaches a private Excel log listing every failed and skipped row."""
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Import Log"
		ws.append(["Row Index", "Target DocType", "Type", "Reason"])
		for entry in row_log:
			ws.append(
				[entry.get("row"), entry.get("doctype"), entry.get("type", "Failed"), entry.get("reason")]
			)

		stream = io.BytesIO()
		wb.save(stream)
		wb.close()

		self._remove_previous_log()
		from frappe.utils.file_manager import save_file

		file_doc = save_file(
			f"Import_Log_{self.doc.name}.xlsx",
			stream.getvalue(),
			"Smart Data Import",
			self.doc.name,
			is_private=1,
		)
		self.doc.error_file = file_doc.file_url

	def _remove_previous_log(self):
		for name in frappe.get_all(
			"File",
			filters={
				"attached_to_doctype": "Smart Data Import",
				"attached_to_name": self.doc.name,
				"file_name": ("like", "Import_Log_%"),
			},
			pluck="name",
		):
			try:
				frappe.delete_doc("File", name, force=True, ignore_permissions=True)
			except Exception:
				pass

	# -------------------------------------------------------------------------
	# 4. ROLLBACK MANIFEST
	# -------------------------------------------------------------------------

	def _record_in_manifest(self, action, target_doctype, name):
		"""Remembers one written record so the import can be undone later.

		`action` is "I" for an inserted record (deletable on rollback) or "U" for an
		updated one (its previous values are not recoverable).
		"""
		self._manifest_buffer.append((action, target_doctype, name))

	def _flush_manifest(self):
		"""Appends the buffered entries to the manifest file on disk.

		Written incrementally and never held in memory, so a manifest for millions of
		rows costs nothing and survives a crash mid-import.
		"""
		if not self._manifest_buffer:
			return
		path = manifest_path(self.doc.name)
		os.makedirs(os.path.dirname(path), exist_ok=True)
		with open(path, "a", encoding="utf-8", newline="") as f:
			csv.writer(f).writerows(self._manifest_buffer)
		self._manifest_buffer = []

	def _attach_manifest(self):
		"""Publishes the manifest as a private attachment so it is visible and
		downloadable, and stores its URL on the document."""
		if not self.supports("rollback_file"):
			return

		path = manifest_path(self.doc.name)
		if not os.path.exists(path):
			return

		file_url = f"/private/files/{os.path.basename(path)}"
		if getattr(self.doc, "rollback_file", None) == file_url and frappe.db.exists(
			"File", {"file_url": file_url}
		):
			return

		if not frappe.db.exists("File", {"file_url": file_url}):
			frappe.get_doc(
				{
					"doctype": "File",
					"file_name": os.path.basename(path),
					"file_url": file_url,
					"is_private": 1,
					"attached_to_doctype": "Smart Data Import",
					"attached_to_name": self.doc.name,
				}
			).insert(ignore_permissions=True)
		self.doc.rollback_file = file_url

	# -------------------------------------------------------------------------
	# 5. ROLLBACK EXECUTION
	# -------------------------------------------------------------------------

	def execute_rollback(self, force=False):
		"""Deletes every record this import created, newest first.

		Newest-first is what makes it safe: child rows, later tiers and tree children
		go before the parents they point at. Records that cannot be deleted (still
		linked elsewhere) are retried in further passes and then reported.
		"""
		start_time = time.time()
		entries = read_manifest(self.doc.name)
		inserted = [e for e in entries if e[0] == "I"]
		updated_count = sum(1 for e in entries if e[0] == "U")

		self.doc.status = "Rolling Back"
		self.doc.progress_percent = 0
		self.doc.rolled_back_records = 0
		self.doc.save(ignore_permissions=True)
		frappe.db.commit()

		# Fail before deleting anything if a DocType may not be deleted from.
		for target_doctype in {e[1] for e in inserted}:
			frappe.has_permission(target_doctype, "delete", throw=True)

		total = len(inserted)
		self.publish_progress(_("Rolling back {0} records...").format(total))

		# Hierarchies link both ways (a child points at its parent, and the parent's
		# own validations refuse to go while a child exists), so no deletion order
		# alone can break the pair. Detach the self-references first.
		self._unlink_self_references(inserted)

		pending = list(reversed(inserted))
		deleted = 0
		problems = {}

		for _pass_no in range(3):
			retry = []
			for _action, target_doctype, name in pending:
				try:
					if self._delete_one(target_doctype, name, force):
						deleted += 1
						problems.pop((target_doctype, name), None)
					frappe.db.commit()
				except Exception as e:
					frappe.db.rollback()
					problems[(target_doctype, name)] = self._clean_error(e)
					retry.append((_action, target_doctype, name))

				if total and deleted % 200 == 0:
					self._publish_rollback_progress(deleted, total)

			if not retry or len(retry) == len(pending):
				pending = retry
				break
			pending = retry

		row_log = [
			{
				"row": 0,
				"doctype": target_doctype,
				"type": "Rollback",
				"reason": _("Could not delete {0}: {1}").format(name, reason),
			}
			for (target_doctype, name), reason in problems.items()
		]

		# Keep only what is still there, so a second rollback attempt retries just those.
		rewrite_manifest(self.doc.name, pending)

		self.doc.reload()
		self.doc.rolled_back_records = deleted
		self.doc.execution_time_seconds = flt(time.time() - start_time, 2)
		self.doc.status = "Rolled Back"
		self.doc.progress_percent = 100
		self.doc.error_log = ""
		self._store_error_rows(row_log)
		if row_log:
			self._generate_failed_rows_excel(row_log)
			self.doc.error_log = "\n".join(entry["reason"] for entry in row_log[:200])
		if updated_count:
			self.doc.error_log = (
				_("{0} record(s) were updated by this import — updates cannot be undone automatically.").format(
					updated_count
				)
				+ ("\n" + self.doc.error_log if self.doc.error_log else "")
			)
		if not pending:
			self._detach_manifest()
		else:
			self._attach_manifest()
		self.doc.save(ignore_permissions=True)
		frappe.db.commit()

		self.publish_progress(
			_("Rollback finished: {0} of {1} records deleted, {2} could not be deleted.").format(
				deleted, total, len(pending)
			),
			100,
		)
		return {"deleted": deleted, "total": total, "failed": len(pending), "updated": updated_count}

	def _unlink_self_references(self, inserted):
		"""Clears self-referencing Link fields (parent_task, parent_item_group, ...)
		on the records that are about to be deleted.

		Tree DocTypes are saved through the document so the nested set is updated;
		flat ones are written directly. Only records whose field is actually set are
		touched, and records outside this import are never modified — a link from
		outside must still block the delete.
		"""
		by_doctype = defaultdict(list)
		for _action, target_doctype, name in inserted:
			by_doctype[target_doctype].append(name)

		for target_doctype, names in by_doctype.items():
			meta = frappe.get_meta(target_doctype)
			self_fields = [
				f.fieldname
				for f in meta.fields
				if f.fieldtype == "Link" and f.options == target_doctype
			]
			if not self_fields:
				continue

			is_tree = bool(getattr(meta, "is_tree", 0)) or (
				meta.has_field("lft") and meta.has_field("rgt")
			)

			for chunk_start in range(0, len(names), 500):
				chunk = names[chunk_start : chunk_start + 500]
				rows = frappe.get_all(
					target_doctype,
					filters={"name": ("in", chunk)},
					fields=["name", *self_fields],
				)
				for row in rows:
					if not any(row.get(f) for f in self_fields):
						continue
					try:
						if is_tree:
							doc = frappe.get_doc(target_doctype, row["name"])
							for fieldname in self_fields:
								doc.set(fieldname, None)
							doc.flags.ignore_permissions = True
							doc.flags.ignore_links = True
							doc.save(ignore_permissions=True)
						else:
							frappe.db.set_value(
								target_doctype,
								row["name"],
								{f: None for f in self_fields if row.get(f)},
								update_modified=False,
							)
					except Exception:
						# Fall back to a direct write; the record is being deleted anyway.
						frappe.db.rollback()
						frappe.db.set_value(
							target_doctype,
							row["name"],
							{f: None for f in self_fields if row.get(f)},
							update_modified=False,
						)
				frappe.db.commit()

	def _delete_one(self, target_doctype, name, force):
		"""Deletes a single record, cancelling it first if it was submitted."""
		if not frappe.db.exists(target_doctype, name):
			return True  # already gone — nothing to undo

		if cint(frappe.db.get_value(target_doctype, name, "docstatus")) == 1:
			doc = frappe.get_doc(target_doctype, name)
			doc.flags.ignore_permissions = True
			doc.cancel()

		frappe.delete_doc(
			target_doctype,
			name,
			force=1 if force else 0,
			ignore_permissions=True,
			ignore_missing=True,
			delete_permanently=True,
		)
		return True

	def _publish_rollback_progress(self, deleted, total):
		percent = flt((deleted / total) * 100, 1) if total else 100
		frappe.db.set_value(
			"Smart Data Import",
			self.doc.name,
			self.supported_updates({"rolled_back_records": deleted, "progress_percent": percent}),
			update_modified=False,
		)
		frappe.db.commit()
		self.doc.progress_percent = percent
		self.doc.rolled_back_records = deleted
		self.publish_progress(_("Deleted {0} of {1} records...").format(deleted, total), percent)

	def _detach_manifest(self):
		"""Removes the manifest once everything it listed is gone."""
		for name in frappe.get_all(
			"File",
			filters={
				"attached_to_doctype": "Smart Data Import",
				"attached_to_name": self.doc.name,
				"file_name": ("like", "Rollback_%"),
			},
			pluck="name",
		):
			try:
				frappe.delete_doc("File", name, force=True, ignore_permissions=True)
			except Exception:
				pass
		path = manifest_path(self.doc.name)
		if os.path.exists(path):
			os.remove(path)
		self.doc.rollback_file = None


# -------------------------------------------------------------------------
# ROLLBACK MANIFEST HELPERS
# -------------------------------------------------------------------------


def manifest_path(doc_name):
	"""Deterministic private path of an import's rollback manifest."""
	return frappe.get_site_path("private", "files", f"Rollback_{doc_name}.csv")


def read_manifest(doc_name):
	"""Returns [(action, doctype, name), ...] for everything this import wrote."""
	path = manifest_path(doc_name)
	if not os.path.exists(path):
		return []

	entries = []
	with open(path, encoding="utf-8", newline="") as f:
		for row in csv.reader(f):
			if len(row) >= 3 and row[0] in ("I", "U"):
				entries.append((row[0], row[1], row[2]))
	return entries


def rewrite_manifest(doc_name, entries):
	"""Replaces the manifest with `entries` (or deletes it when nothing is left)."""
	path = manifest_path(doc_name)
	if not entries:
		if os.path.exists(path):
			os.remove(path)
		return

	with open(path, "w", encoding="utf-8", newline="") as f:
		csv.writer(f).writerows(entries)


# -------------------------------------------------------------------------
# TEMPLATE GENERATION
# -------------------------------------------------------------------------


def _template_columns(meta):
	"""Splits a DocType's real, fillable fields into (mandatory, optional)."""
	mandatory, optional = [], []
	for field in meta.fields:
		if field.fieldtype in LAYOUT_FIELDTYPES or field.fieldtype == "Table":
			continue
		if field.read_only or getattr(field, "is_virtual", 0):
			continue
		(mandatory if field.reqd else optional).append(field)
	return mandatory, optional


def _unique_sheet_name(name, used):
	"""Sanitizes a DocType name into a valid, unique Excel sheet name (<=31 chars,
	no \\/*?:[] characters)."""
	base = re.sub(r"[\\/*?:\[\]]", "", name)[:31] or "Sheet"
	candidate = base
	suffix_no = 1
	while candidate in used:
		suffix = f" ({suffix_no})"
		candidate = base[: 31 - len(suffix)] + suffix
		suffix_no += 1
	return candidate


def _write_template_sheet(ws, columns):
	"""Fills one data-entry sheet with the styled header row for `columns`."""
	headers = [f.label or f.fieldname for f in columns]
	ws.append(headers)
	for idx, field in enumerate(columns, start=1):
		cell = ws.cell(row=1, column=idx)
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill("solid", fgColor="C0392B" if field.reqd else "34495E")
		ws.column_dimensions[cell.column_letter].width = max(14, min(38, len(str(cell.value)) + 6))
	ws.freeze_panes = "A2"


def _field_guide_row(field):
	options = ""
	if field.fieldtype == "Link":
		options = _("Existing {0} (name or title)").format(field.options)
	elif field.fieldtype == "Select" and field.options:
		options = ", ".join(field.options.split("\n")[:12])
	return [field.label or field.fieldname, field.fieldname, field.fieldtype, "Yes" if field.reqd else "No", options]


def build_template_workbook(target_doctype, include_optional=True):
	"""Builds a standalone .xlsx template for one DocType: sheet 1 has the headers
	to fill, sheet 2 documents every column (type, mandatory, link target, options)."""
	meta = frappe.get_meta(target_doctype)
	mandatory, optional = _template_columns(meta)
	columns = mandatory + (optional if include_optional else [])

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = _unique_sheet_name(target_doctype, set())
	_write_template_sheet(ws, columns)

	guide = wb.create_sheet("Field Guide")
	guide.append(["Column Header", "Fieldname", "Type", "Mandatory", "Links To / Options"])
	for field in columns:
		guide.append(_field_guide_row(field))
	for column_cells in guide.columns:
		guide.column_dimensions[column_cells[0].column_letter].width = 32
	for cell in guide[1]:
		cell.font = Font(bold=True)

	stream = io.BytesIO()
	wb.save(stream)
	wb.close()
	return stream.getvalue(), len(mandatory), len(columns)


def build_combined_template_workbook(target_doctypes, include_optional=True):
	"""Builds one .xlsx with one data-entry sheet per DocType, plus a single
	combined Field Guide sheet (with a DocType column) documenting all of them.

	Returns (content_bytes, stats) where stats is
	[{"doctype", "mandatory_columns", "total_columns"}, ...] in input order.
	"""
	wb = openpyxl.Workbook()
	wb.remove(wb.active)

	used_sheet_names = set()
	guide_rows = []
	stats = []

	for target_doctype in target_doctypes:
		meta = frappe.get_meta(target_doctype)
		mandatory, optional = _template_columns(meta)
		columns = mandatory + (optional if include_optional else [])

		ws = wb.create_sheet(_unique_sheet_name(target_doctype, used_sheet_names))
		used_sheet_names.add(ws.title)
		_write_template_sheet(ws, columns)

		for field in columns:
			guide_rows.append([target_doctype, *_field_guide_row(field)])

		stats.append(
			{"doctype": target_doctype, "mandatory_columns": len(mandatory), "total_columns": len(columns)}
		)

	guide = wb.create_sheet("Field Guide")
	guide.append(["DocType", "Column Header", "Fieldname", "Type", "Mandatory", "Links To / Options"])
	for row in guide_rows:
		guide.append(row)
	for column_cells in guide.columns:
		guide.column_dimensions[column_cells[0].column_letter].width = 28
	for cell in guide[1]:
		cell.font = Font(bold=True)

	stream = io.BytesIO()
	wb.save(stream)
	wb.close()
	return stream.getvalue(), stats


def build_template_zip(target_doctypes, include_optional=True):
	"""Builds a .zip containing one standalone .xlsx template per DocType.

	Returns (zip_bytes, stats), same stats shape as build_combined_template_workbook.
	"""
	import zipfile

	stats = []
	used_names = set()
	buffer = io.BytesIO()
	with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
		for target_doctype in target_doctypes:
			content, mandatory_count, total_columns = build_template_workbook(
				target_doctype, include_optional=include_optional
			)
			entry_name = f"Template_{target_doctype.replace(' ', '_')}.xlsx"
			suffix_no = 1
			while entry_name in used_names:
				entry_name = f"Template_{target_doctype.replace(' ', '_')}_{suffix_no}.xlsx"
				suffix_no += 1
			used_names.add(entry_name)
			zf.writestr(entry_name, content)
			stats.append(
				{"doctype": target_doctype, "mandatory_columns": mandatory_count, "total_columns": total_columns}
			)

	return buffer.getvalue(), stats


# -------------------------------------------------------------------------
# FRAPPE BACKGROUND JOB WRAPPER & API METHODS
# -------------------------------------------------------------------------


def _get_import_doc(doc_name, ptype="write"):
	doc = frappe.get_doc("Smart Data Import", doc_name)
	doc.check_permission(ptype)
	return doc


def required_permissions_for_mode(import_type):
	"""Which permissions an import mode actually needs. Single source of truth:
	"Insert and Update" does both per row, so it needs both — deriving this in more
	than one place is how a write-only role once slipped through and created records.
	"""
	permissions = []
	if import_type in ("Insert New Records", "Insert and Update"):
		permissions.append("create")
	if import_type in ("Update Existing Records", "Insert and Update"):
		permissions.append("write")
	return permissions


def assert_can_import_into(target_doctype, import_type, user=None):
	user = user or frappe.session.user
	if target_doctype in SECURITY_CRITICAL_DOCTYPES:
		frappe.throw(
			_("You do not have permission to import into {0}.").format(target_doctype),
			frappe.PermissionError,
		)
	for ptype in required_permissions_for_mode(import_type):
		if not frappe.has_permission(target_doctype, ptype, user=user):
			frappe.throw(
				_("You do not have permission to {0} {1} records.").format(ptype, target_doctype),
				frappe.PermissionError,
			)


def _assert_target_doctypes_allowed(doc):
	"""System Managers keep using this exactly as before — this only restricts
	everyone else (e.g. a self-service portal user), since the engine itself writes
	with ignore_permissions=True regardless of who started the job."""
	if "System Manager" in frappe.get_roles(frappe.session.user):
		return

	for target_dt in {dep.doctype_name for dep in doc.dependencies if dep.doctype_name}:
		assert_can_import_into(target_dt, doc.import_type)


@frappe.whitelist()
def analyze_smart_import(doc_name):
	engine = SmartImportEngine(_get_import_doc(doc_name))
	engine.analyze_files_and_build_graph()
	return frappe.get_doc("Smart Data Import", doc_name)


@frappe.whitelist()
def start_smart_import(doc_name):
	doc = _get_import_doc(doc_name)

	if doc.status == "Processing":
		return {"status": "error", "message": _("This import is already running.")}
	if not doc.files:
		frappe.throw(_("Attach at least one Excel or CSV file before starting the import."))

	# One click is enough: analyze automatically if the graph is missing or stale.
	if doc.status in ("Pending", "Analyzing") or not doc.dependencies:
		SmartImportEngine(doc).analyze_files_and_build_graph()
		doc.reload()

	blocked = [row.idx for row in doc.files if row.status == "Failed" and not cint(row.total_rows)]
	if not doc.dependencies:
		frappe.throw(
			_("Nothing to import. Check the file rows for errors: {0}").format(
				", ".join(str(i) for i in blocked) or _("no readable rows found")
			)
		)

	_assert_target_doctypes_allowed(doc)

	frappe.enqueue(
		"smart_data_import.smart_import_engine.run_async_import",
		queue="long",
		timeout=7200,
		doc_name=doc_name,
	)
	frappe.db.set_value("Smart Data Import", doc_name, "status", "Processing", update_modified=False)
	frappe.db.commit()
	return {
		"status": "queued",
		"message": _("Import queued in the background for {0} rows.").format(doc.total_records),
	}


@frappe.whitelist()
def cancel_import(doc_name):
	"""Flags a running import to stop. The background job itself checks this flag
	between batches (see SmartImportEngine._raise_if_cancelled) and stops cleanly —
	whatever was already imported at that point is kept, nothing is rolled back."""
	doc = _get_import_doc(doc_name)
	if doc.status != "Processing":
		return {"status": "error", "message": _("Only a running import can be cancelled.")}

	# The background job saves this same row on nearly every batch, so this write
	# can collide with it — retry rather than surface a transient deadlock to the user.
	for attempt in range(5):
		try:
			frappe.db.set_value("Smart Data Import", doc_name, "status", "Cancelled", update_modified=False)
			frappe.db.commit()
			break
		except frappe.QueryDeadlockError:
			frappe.db.rollback()
			if attempt == 4:
				raise
			time.sleep(0.2)

	return {
		"status": "cancelled",
		"message": _("Cancelling — this will stop after the batch currently in progress."),
	}


@frappe.whitelist()
def reset_import(doc_name):
	"""Clears counters and statuses so a finished or failed import can be re-run.

	The rollback manifest is deliberately kept: the records created by the previous
	run still exist, so it must stay possible to undo them.
	"""
	doc = _get_import_doc(doc_name)
	if doc.status in ("Processing", "Rolling Back"):
		frappe.throw(_("Cannot reset while the import is running."))

	doc.status = "Pending"
	doc.imported_records = 0
	doc.failed_records = 0
	doc.skipped_records = 0
	doc.rolled_back_records = 0
	doc.progress_percent = 0
	doc.execution_time_seconds = 0
	doc.error_log = ""
	doc.error_file = None
	doc.errors = []
	for row in doc.files:
		row.status = "Pending"
		row.error_log = ""
	for row in doc.dependencies:
		row.status = "Pending"
		row.processed_count = 0
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"status": "reset", "message": _("Import reset. Re-analyze and start again when ready.")}


@frappe.whitelist()
def get_import_preview(doc_name, sample_size=5):
	"""Returns, per attached file, exactly how columns will be mapped plus a few
	sample rows — so mistakes are visible before any record is written."""
	doc = _get_import_doc(doc_name, "read")
	engine = SmartImportEngine(doc)
	sample_size = min(20, max(1, cint(sample_size) or 5))
	result = []

	for row in doc.files:
		entry = {
			"idx": row.idx,
			"file": row.file,
			"doctype": row.doctype_name,
			"sheet": row.sheet_name,
			"total_rows": cint(row.total_rows),
			"error": None,
			"mapped": {},
			"unmapped": [],
			"missing_mandatory": [],
			"mandatory_fieldnames": [],
			"headers": [],
			"samples": [],
		}

		file_path, error = resolve_file_path(row.file)
		if error:
			entry["error"] = error
			result.append(entry)
			continue

		if not row.doctype_name:
			entry["error"] = _("Target DocType not set for this file.")
			result.append(entry)
			continue

		generator = engine._get_row_generator(file_path, row.sheet_name)
		headers = SmartImportEngine._clean_header_row(next(generator, []) or [])
		samples = []
		for values in generator:
			samples.append(["" if v is None else str(v) for v in values])
			if len(samples) >= sample_size:
				break

		meta = frappe.get_meta(row.doctype_name)
		column_map, defaults = engine.rules_for(row.doctype_name)
		mapping = map_columns(meta, headers, column_map)

		entry["headers"] = headers
		entry["samples"] = samples
		entry["mapped"] = mapping["mapped"]
		entry["unmapped"] = mapping["unmapped"]
		entry["missing_mandatory"] = missing_mandatory_fields(meta, mapping["mapped"].values(), defaults)
		entry["mandatory_fieldnames"] = sorted(mandatory_fieldnames(meta))
		entry["defaults"] = defaults
		result.append(entry)

	return {"files": result, "import_type": doc.import_type, "status": doc.status}


def _parse_template_doctypes(target_doctype):
	"""Accepts either a single DocType name or a JSON-encoded list of names
	(a Dialog MultiSelectList sends the latter), de-duplicated in order."""
	if isinstance(target_doctype, list):
		items = target_doctype
	else:
		text = (target_doctype or "").strip()
		if text.startswith("["):
			try:
				items = json.loads(text)
			except Exception:
				items = [text]
		else:
			items = [text]

	seen = []
	for item in items:
		item = (item or "").strip()
		if item and item not in seen:
			seen.append(item)
	if not seen:
		frappe.throw(_("Select at least one DocType."))
	return seen


@frappe.whitelist()
def download_import_template(target_doctype, include_optional=1, mode="single"):
	"""Generates ready-to-fill Excel template(s) and returns the download URL.

	`target_doctype` may be a single DocType name or a JSON-encoded list of names.
	`mode` only matters when more than one DocType is given: "single" builds one
	.xlsx with one sheet per DocType; "separate" builds a .zip with one
	standalone .xlsx per DocType.
	"""
	doctypes = _parse_template_doctypes(target_doctype)
	for dt in doctypes:
		if not frappe.db.exists("DocType", dt):
			frappe.throw(_("DocType {0} does not exist.").format(dt))
		frappe.has_permission(dt, "create", throw=True)

	include_optional = cint(include_optional)

	if len(doctypes) == 1:
		content, mandatory_count, total_columns = build_template_workbook(
			doctypes[0], include_optional=include_optional
		)
		stats = [{"doctype": doctypes[0], "mandatory_columns": mandatory_count, "total_columns": total_columns}]
		filename = f"Template_{doctypes[0].replace(' ', '_')}.xlsx"
	elif mode == "separate":
		content, stats = build_template_zip(doctypes, include_optional=include_optional)
		filename = "Import_Templates.zip"
	else:
		content, stats = build_combined_template_workbook(doctypes, include_optional=include_optional)
		filename = "Import_Templates.xlsx"

	from frappe.utils.file_manager import save_file

	file_doc = save_file(filename, content, None, None, is_private=1)
	return {
		"file_url": file_doc.file_url,
		"file_name": file_doc.file_name,
		"stats": stats,
	}


def error_signature(reason):
	"""Collapses row-specific details so 496 failures group into a handful of causes.

	"Could not find Customer Group: BH-Vendor" and "...: Marketing -Suppliers" both
	become "Could not find Customer Group: …".
	"""
	text = str(reason or "").strip()
	text = re.sub(r'"[^"]*"', '"…"', text)
	text = re.sub(r"\b[A-Z][A-Z0-9]*-[\w-]*\d[\w-]*\b", "…", text)
	text = re.sub(r"(:\s*)(?!…).+$", r"\1…", text)
	text = re.sub(r"\s+", " ", text)
	return text[:200] or _("Unknown error")


@frappe.whitelist()
def get_error_summary(doc_name, limit=25):
	"""Groups the stored problem rows by cause, most frequent first — this is what
	the "View Errors" dialog shows."""
	doc = _get_import_doc(doc_name, "read")
	limit = min(100, max(1, cint(limit) or 25))

	error_rows = doc.get("errors") or []
	groups = {}
	for row in error_rows:
		key = (row.error_type or "Failed", error_signature(row.reason))
		group = groups.setdefault(
			key,
			{
				"type": key[0],
				"problem": key[1],
				"count": 0,
				"example_row": row.row_index,
				"example_reason": row.reason,
				"doctypes": set(),
			},
		)
		group["count"] += 1
		if row.doctype_name:
			group["doctypes"].add(row.doctype_name)

	ranked = sorted(groups.values(), key=lambda g: (-g["count"], g["problem"]))[:limit]
	for group in ranked:
		group["doctypes"] = ", ".join(sorted(group["doctypes"]))

	return {
		"groups": ranked,
		"shown_rows": len(error_rows),
		"failed": cint(doc.failed_records),
		"skipped": cint(doc.get("skipped_records")),
		"truncated": len(error_rows) >= ERROR_ROWS_IN_UI,
		"log_file": doc.error_file,
		"status": doc.status,
	}


def may_delete_manifest_doctypes(entries):
	"""Whether the caller could actually delete what a rollback would remove.

	Reported alongside the summary so the UI never offers an Undo action that is
	certain to be refused — rollback_import enforces the same rule for real.
	"""
	return all(
		frappe.has_permission(target_doctype, "delete")
		for target_doctype in {e[1] for e in entries if e[0] == "I"}
	)


@frappe.whitelist()
def get_rollback_summary(doc_name):
	"""What an "Undo Import" would delete, per DocType — read-only."""
	doc = _get_import_doc(doc_name, "read")
	entries = read_manifest(doc_name)

	per_doctype = {}
	for action, target_doctype, _name in entries:
		bucket = per_doctype.setdefault(target_doctype, {"inserted": 0, "updated": 0})
		bucket["inserted" if action == "I" else "updated"] += 1

	return {
		"deletable": sum(1 for e in entries if e[0] == "I"),
		"updated": sum(1 for e in entries if e[0] == "U"),
		"doctypes": [
			{"doctype": dt, "inserted": v["inserted"], "updated": v["updated"]}
			for dt, v in sorted(per_doctype.items())
		],
		"status": doc.status,
		"can_rollback": (
			bool(entries)
			and doc.status not in ("Processing", "Rolling Back")
			and may_delete_manifest_doctypes(entries)
		),
	}


@frappe.whitelist()
def rollback_import(doc_name, force=0):
	"""Queues deletion of every record this import created.

	`force` skips Frappe's link validations — it can break documents that reference
	the deleted records, so it stays opt-in.
	"""
	doc = _get_import_doc(doc_name, "delete")
	if doc.status in ("Processing", "Rolling Back"):
		frappe.throw(_("Wait for the current job to finish before rolling back."))

	entries = read_manifest(doc_name)
	if not entries:
		frappe.throw(
			_("Nothing to roll back — no record of created documents exists for this import.")
		)
	if not any(e[0] == "I" for e in entries):
		frappe.throw(
			_("This import only updated existing records, which cannot be undone automatically.")
		)

	# Explicit message: has_permission(throw=True) raises with an empty string here,
	# which reaches the user as a blank error.
	for target_doctype in {e[1] for e in entries if e[0] == "I"}:
		if not frappe.has_permission(target_doctype, "delete"):
			frappe.throw(
				_("This import cannot be undone: you do not have permission to delete {0} records.").format(
					target_doctype
				),
				frappe.PermissionError,
			)

	frappe.enqueue(
		"smart_data_import.smart_import_engine.run_async_rollback",
		queue="long",
		timeout=7200,
		doc_name=doc_name,
		force=cint(force),
	)
	frappe.db.set_value("Smart Data Import", doc_name, "status", "Rolling Back", update_modified=False)
	frappe.db.commit()
	return {
		"status": "queued",
		"message": _("Rollback queued: deleting {0} records in the background.").format(
			sum(1 for e in entries if e[0] == "I")
		),
	}


@frappe.whitelist()
def get_importable_doctypes(txt=""):
	"""Link-query friendly list of DocTypes that can be imported into."""
	filters = {"istable": 0, "issingle": 0}
	if txt:
		filters["name"] = ("like", f"%{txt}%")
	return frappe.get_all("DocType", filters=filters, pluck="name", order_by="name", limit_page_length=50)


@frappe.whitelist()
def portal_importable_doctypes():
	"""DocTypes the self-service /import page may offer to the current user — only
	ones they can actually create, and never the security-critical ones regardless
	of role (see SECURITY_CRITICAL_DOCTYPES)."""
	candidates = frappe.get_all("DocType", filters={"istable": 0, "issingle": 0}, pluck="name")
	return sorted(
		dt
		for dt in candidates
		if dt not in SECURITY_CRITICAL_DOCTYPES and frappe.has_permission(dt, "create")
	)


PORTAL_TOGGLE_OPTIONS = (
	"ignore_link_errors",
	"ignore_mandatory_errors",
	"stop_on_error",
	"ignore_duplicates",
	"skip_empty_rows",
	"clean_whitespace",
)


def excel_sheet_names(file_path):
	"""Sheet names in a workbook, so one file row can be created per sheet.

	read_file_header_and_count only ever reads a single sheet, so a workbook with
	several sheets needs one row each or the extra sheets are silently ignored.
	"""
	workbook = openpyxl.load_workbook(file_path, read_only=True)
	try:
		return list(workbook.sheetnames)
	finally:
		workbook.close()


def _claim_portal_file(file_url):
	"""Ownership gate for a browser-supplied file_url.

	Uploads arrive standalone (no attached_to_*), so this is the only place
	ownership is checked — without it any logged-in user could pass the file_url of
	someone else's private upload and have it read and re-parented onto their import.
	"""
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	if file_doc.owner != frappe.session.user and not frappe.has_permission("File", "read", doc=file_doc):
		frappe.throw(_("You do not have permission to use this file."), frappe.PermissionError)
	if file_doc.attached_to_name:
		frappe.throw(_("This file is already attached to another record."), frappe.PermissionError)
	return file_doc


def _portal_plan(doc):
	"""Everything the /import_data page needs to show after analysis: what each file
	was detected as, the execution order, and anything the caller may not import."""
	verdicts = {}

	def refusal(target_doctype):
		if target_doctype not in verdicts:
			try:
				assert_can_import_into(target_doctype, doc.import_type)
				verdicts[target_doctype] = ""
			except frappe.PermissionError as e:
				verdicts[target_doctype] = str(e) or _("You may not import into {0}.").format(target_doctype)
		return verdicts[target_doctype]

	files = []
	for row in doc.files:
		files.append(
			{
				"row": row.name,
				"idx": row.idx,
				"file_name": os.path.basename((row.file or "").split("?")[0]),
				"sheet_name": row.sheet_name,
				"doctype_name": row.doctype_name,
				"total_rows": cint(row.total_rows),
				"status": row.status,
				"mapping_summary": row.mapping_summary,
				"error_log": row.error_log,
				"refused": refusal(row.doctype_name) if row.doctype_name else "",
			}
		)

	plan = [
		{
			"tier": cint(dep.execution_tier),
			"doctype": dep.doctype_name,
			"depends_on": dep.depends_on_doctypes,
			"total_count": cint(dep.total_count),
			"self_reference_field": dep.self_reference_field,
		}
		for dep in sorted(doc.dependencies, key=lambda d: cint(d.execution_tier))
	]

	refusals = sorted({f["refused"] for f in files if f["refused"]})
	return {
		"doc_name": doc.name,
		"status": doc.status,
		"import_type": doc.import_type,
		"total_records": cint(doc.total_records),
		"files": files,
		"plan": plan,
		"refusals": refusals,
		"undetected": [f["idx"] for f in files if not f["doctype_name"]],
		"can_start": bool(plan) and not refusals,
	}


def _analyzed_plan(doc):
	SmartImportEngine(doc).analyze_files_and_build_graph()
	doc.reload()
	return _portal_plan(doc)


@frappe.whitelist()
def portal_create_import(files, import_type="Insert New Records", options=None):
	"""Creates and analyzes a multi-file import for /import_data without starting it.

	`files` is a list of uploaded file URLs, or of {file_url, doctype_name} when the
	user has already chosen the target. DocTypes left blank are auto-detected, and
	multi-sheet workbooks are expanded into one row per sheet. Starting is a separate
	call to start_smart_import, which re-checks every detected DocType anyway.
	"""
	files = frappe.parse_json(files) if isinstance(files, str) else files
	if not files:
		frappe.throw(_("Upload at least one file."))
	options = frappe.parse_json(options) if isinstance(options, str) else (options or {})

	doc = frappe.new_doc("Smart Data Import")
	doc.title = _("Import by {0} on {1}").format(frappe.session.user, frappe.utils.now())
	doc.import_type = import_type
	doc.auto_detect_doctype = 1
	# Explicit allowlist, never doc.update(options): the payload comes from the
	# browser and must not be able to set status, owner or any other field.
	for fieldname in PORTAL_TOGGLE_OPTIONS:
		if fieldname in options:
			doc.set(fieldname, cint(options.get(fieldname)))
	if options.get("batch_size"):
		doc.batch_size = cint(options["batch_size"])
	if options.get("filter_rules_json"):
		doc.filter_rules_json = options["filter_rules_json"]

	claimed = []
	for entry in files:
		entry = {"file_url": entry} if isinstance(entry, str) else entry
		file_url = entry.get("file_url")
		chosen = (entry.get("doctype_name") or "").strip()
		if chosen:
			assert_can_import_into(chosen, import_type)

		claimed.append(_claim_portal_file(file_url))

		sheets = []
		file_path, error = resolve_file_path(file_url)
		if not error and file_path.lower().endswith(EXCEL_EXTENSIONS):
			try:
				sheets = excel_sheet_names(file_path)
			except Exception:
				sheets = []

		# One row per sheet only when the workbook actually has several; a single
		# sheet keeps sheet_name blank so the engine picks it on its own.
		if len(sheets) > 1:
			for sheet_name in sheets:
				doc.append("files", {"file": file_url, "doctype_name": chosen, "sheet_name": sheet_name})
		else:
			doc.append("files", {"file": file_url, "doctype_name": chosen})

	doc.insert()  # "All" role + if_owner grants this — not a System Manager-only action.

	for file_doc in claimed:
		file_doc.attached_to_doctype = "Smart Data Import"
		file_doc.attached_to_name = doc.name
		file_doc.save(ignore_permissions=True)

	return _analyzed_plan(doc)


@frappe.whitelist()
def portal_set_file_doctypes(doc_name, assignments):
	"""Applies the user's corrections to auto-detection and re-analyzes.

	Re-analysis is safe: detection only fills blank DocTypes, so a choice made here
	is never overwritten.
	"""
	doc = _get_import_doc(doc_name)
	assignments = frappe.parse_json(assignments) if isinstance(assignments, str) else (assignments or {})

	for row_name, target_doctype in assignments.items():
		target_doctype = (target_doctype or "").strip()
		if target_doctype:
			assert_can_import_into(target_doctype, doc.import_type)
		for row in doc.files:
			if row.name == row_name:
				row.doctype_name = target_doctype
				break

	doc.save()
	return _analyzed_plan(doc)


@frappe.whitelist()
def portal_remove_file(doc_name, row_name):
	"""Drops one file/sheet row — used to discard sheets that aren't data
	(cover pages, notes) instead of leaving them reported as undetected."""
	doc = _get_import_doc(doc_name)
	doc.files = [row for row in doc.files if row.name != row_name]
	for position, row in enumerate(doc.files, start=1):
		row.idx = position
	doc.save()
	return _analyzed_plan(doc)


@frappe.whitelist()
def portal_plan(doc_name):
	"""Current analysis result without re-reading the files."""
	return _portal_plan(_get_import_doc(doc_name, "read"))


@frappe.whitelist()
def portal_my_imports(limit=8):
	"""The current user's own recent imports, for the history panel on /import_data."""
	return frappe.get_all(
		"Smart Data Import",
		filters={"owner": frappe.session.user},
		fields=[
			"name",
			"title",
			"status",
			"total_records",
			"imported_records",
			"failed_records",
			"modified",
		],
		order_by="modified desc",
		limit_page_length=min(50, max(1, cint(limit) or 8)),
	)


@frappe.whitelist()
def portal_get_status(doc_name):
	"""Polling endpoint for the portal page — avoids depending on the realtime/
	socketio stack being reachable from a plain website page."""
	doc = _get_import_doc(doc_name, "read")
	return {
		"doc_name": doc.name,
		# A single import can span several DocTypes, so report all of them.
		"doctypes": sorted({row.doctype_name for row in doc.files if row.doctype_name}),
		"import_type": doc.import_type,
		"status": doc.status,
		"progress": doc.progress_percent,
		"total": cint(doc.total_records),
		"imported": cint(doc.imported_records),
		"failed": cint(doc.failed_records),
		"skipped": cint(getattr(doc, "skipped_records", 0)),
		"rolled_back": cint(getattr(doc, "rolled_back_records", 0)),
		"seconds": doc.execution_time_seconds,
		"error_file": doc.error_file,
		# Read the manifest only once the run has settled — during Processing this is
		# polled every couple of seconds and the manifest can be very large.
		"can_rollback": bool(
			doc.rollback_file
			and doc.status not in ("Processing", "Rolling Back")
			and may_delete_manifest_doctypes(read_manifest(doc.name))
		),
	}


def run_async_import(doc_name):
	"""Background worker entry point. Any crash must leave the document in a
	recoverable state, otherwise the UI stays stuck on 'Processing' forever."""
	engine = None
	try:
		engine = SmartImportEngine(doc_name)
		engine.execute_import()
	except Exception:
		traceback_text = frappe.get_traceback()
		frappe.db.rollback()
		frappe.log_error(title=f"Smart Data Import failed: {doc_name}", message=traceback_text)
		frappe.db.set_value(
			"Smart Data Import",
			doc_name,
			{
				"status": "Failed",
				"error_log": traceback_text[-5000:],
				"progress_percent": 100,
			},
			update_modified=False,
		)
		frappe.db.commit()
		if engine:
			engine.doc.status = "Failed"
			engine.publish_progress(_("Import failed. See the Detailed Error Log."), 100)
		raise


def run_async_rollback(doc_name, force=0):
	"""Background worker entry point for rollback. Like the import, a crash must not
	leave the document stuck on 'Rolling Back'."""
	engine = None
	try:
		engine = SmartImportEngine(doc_name)
		return engine.execute_rollback(force=bool(cint(force)))
	except Exception:
		traceback_text = frappe.get_traceback()
		frappe.db.rollback()
		frappe.log_error(title=f"Smart Data Import rollback failed: {doc_name}", message=traceback_text)
		frappe.db.set_value(
			"Smart Data Import",
			doc_name,
			{
				"status": "Failed",
				"error_log": traceback_text[-5000:],
				"progress_percent": 100,
			},
			update_modified=False,
		)
		frappe.db.commit()
		if engine:
			engine.doc.status = "Failed"
			engine.publish_progress(_("Rollback failed. See the Detailed Error Log."), 100)
		raise
